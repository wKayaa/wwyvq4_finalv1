#!/usr/bin/env python3
"""
📊 WWYVQ Framework v2.1 - Export and Reporting Module
Ultra-Organized Architecture - Professional Data Export & Reporting

Features:
- Multi-format data export (JSON, CSV, YAML, HTML)
- Professional report generation
- Session backup and archiving
- Result organization and categorization
- Statistical analysis
- Timeline tracking
- Rich HTML reports with charts
"""

import json
import csv
import yaml
import os
from pathlib import Path
from typing import Dict, List, Optional, Any, Union
from dataclasses import dataclass, asdict
from enum import Enum
from datetime import datetime, timedelta
import hashlib
import zipfile
import tempfile


class ExportFormat(Enum):
    """Export formats"""
    JSON = "json"
    CSV = "csv"
    YAML = "yaml"
    HTML = "html"
    XML = "xml"
    EXCEL = "xlsx"


class ReportType(Enum):
    """Report types"""
    SUMMARY = "summary"
    DETAILED = "detailed"
    EXECUTIVE = "executive"
    TECHNICAL = "technical"
    CREDENTIALS = "credentials"
    STATISTICS = "statistics"


@dataclass
class ExportResult:
    """Export operation result"""
    export_id: str
    format: ExportFormat
    file_path: str
    file_size: int
    records_count: int
    export_time: float
    checksum: str
    metadata: Dict[str, Any] = None
    
    def __post_init__(self):
        if self.metadata is None:
            self.metadata = {}


class ExporterModule:
    """
    Professional data export and reporting module
    
    Features:
    - Multi-format export capabilities
    - Automated report generation
    - Data archival and backup
    - Statistical analysis
    """
    
    def __init__(self, config_manager, logger, engine):
        """Initialize exporter module"""
        self.config_manager = config_manager
        self.logger = logger
        self.engine = engine
        self.config = config_manager.get_config().exporter
        
        # Export directory structure
        self.base_export_dir = Path(self.config.export_path)
        self.exports_dir = self.base_export_dir / "exports"
        self.reports_dir = self.base_export_dir / "reports"
        self.archives_dir = self.base_export_dir / "archives"
        
        # Create directories
        for directory in [self.exports_dir, self.reports_dir, self.archives_dir]:
            directory.mkdir(parents=True, exist_ok=True)
        
        # Statistics
        self.stats = {
            'total_exports': 0,
            'total_reports': 0,
            'total_archives': 0,
            'by_format': {},
            'by_type': {},
            'total_size_bytes': 0
        }
        
        self.logger.info("📊 Data Export & Reporting Module initialized")
    
    async def export_data(self, data: Any, format: ExportFormat, 
                         filename: Optional[str] = None) -> ExportResult:
        """Export data in specified format"""
        try:
            import time
            start_time = time.time()
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"wwyvq_export_{timestamp}.{format.value}"
            
            # Ensure correct extension
            if not filename.endswith(f".{format.value}"):
                filename = f"{filename}.{format.value}"
            
            file_path = self.exports_dir / filename
            
            # Export data based on format
            if format == ExportFormat.JSON:
                await self._export_json(data, file_path)
            elif format == ExportFormat.CSV:
                await self._export_csv(data, file_path)
            elif format == ExportFormat.YAML:
                await self._export_yaml(data, file_path)
            elif format == ExportFormat.HTML:
                await self._export_html(data, file_path)
            elif format == ExportFormat.XML:
                await self._export_xml(data, file_path)
            elif format == ExportFormat.EXCEL:
                await self._export_excel(data, file_path)
            else:
                raise ValueError(f"Unsupported export format: {format}")
            
            # Calculate metrics
            file_size = file_path.stat().st_size
            records_count = self._count_records(data)
            export_time = time.time() - start_time
            checksum = self._calculate_checksum(file_path)
            
            # Create export result
            export_id = hashlib.md5(f"{filename}{time.time()}".encode()).hexdigest()[:8]
            
            result = ExportResult(
                export_id=export_id,
                format=format,
                file_path=str(file_path),
                file_size=file_size,
                records_count=records_count,
                export_time=export_time,
                checksum=checksum,
                metadata={
                    'created_at': datetime.now().isoformat(),
                    'source': 'wwyvq_v2.1'
                }
            )
            
            # Update statistics
            self.stats['total_exports'] += 1
            self.stats['total_size_bytes'] += file_size
            
            format_name = format.value
            if format_name not in self.stats['by_format']:
                self.stats['by_format'][format_name] = 0
            self.stats['by_format'][format_name] += 1
            
            self.logger.info(f"📊 Data exported to {file_path} ({file_size} bytes)")
            return result
            
        except Exception as e:
            self.logger.error(f"Failed to export data: {e}")
            raise
    
    async def generate_report(self, report_type: ReportType, data: Dict[str, Any],
                            template: Optional[str] = None) -> str:
        """Generate professional report"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_filename = f"wwyvq_report_{report_type.value}_{timestamp}.html"
            report_path = self.reports_dir / report_filename
            
            # Generate report based on type
            if report_type == ReportType.SUMMARY:
                html_content = await self._generate_summary_report(data)
            elif report_type == ReportType.DETAILED:
                html_content = await self._generate_detailed_report(data)
            elif report_type == ReportType.EXECUTIVE:
                html_content = await self._generate_executive_report(data)
            elif report_type == ReportType.TECHNICAL:
                html_content = await self._generate_technical_report(data)
            elif report_type == ReportType.CREDENTIALS:
                html_content = await self._generate_credentials_report(data)
            elif report_type == ReportType.STATISTICS:
                html_content = await self._generate_statistics_report(data)
            else:
                raise ValueError(f"Unsupported report type: {report_type}")
            
            # Write report
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            # Update statistics
            self.stats['total_reports'] += 1
            
            type_name = report_type.value
            if type_name not in self.stats['by_type']:
                self.stats['by_type'][type_name] = 0
            self.stats['by_type'][type_name] += 1
            
            self.logger.info(f"📋 Report generated: {report_path}")
            return str(report_path)
            
        except Exception as e:
            self.logger.error(f"Failed to generate report: {e}")
            raise
    
    async def archive_session(self, session_id: str, include_logs: bool = True) -> str:
        """Archive complete session data"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            archive_filename = f"session_{session_id}_{timestamp}.zip"
            archive_path = self.archives_dir / archive_filename
            
            with zipfile.ZipFile(archive_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # Add session metadata
                session_data = await self._get_session_data(session_id)
                if session_data:
                    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
                        json.dump(session_data, tmp, indent=2)
                        zipf.write(tmp.name, f"session_{session_id}.json")
                        os.unlink(tmp.name)
                
                # Add results
                results_data = await self._get_session_results(session_id)
                if results_data:
                    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
                        json.dump(results_data, tmp, indent=2)
                        zipf.write(tmp.name, f"results_{session_id}.json")
                        os.unlink(tmp.name)
                
                # Add logs if requested
                if include_logs:
                    logs_data = await self._get_session_logs(session_id)
                    if logs_data:
                        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
                            json.dump(logs_data, tmp, indent=2)
                            zipf.write(tmp.name, f"logs_{session_id}.json")
                            os.unlink(tmp.name)
                
                # Add generated reports
                await self._add_session_reports(zipf, session_id)
            
            # Update statistics
            self.stats['total_archives'] += 1
            archive_size = archive_path.stat().st_size
            self.stats['total_size_bytes'] += archive_size
            
            self.logger.info(f"📦 Session archived: {archive_path} ({archive_size} bytes)")
            return str(archive_path)
            
        except Exception as e:
            self.logger.error(f"Failed to archive session {session_id}: {e}")
            raise
    
    async def _export_json(self, data: Any, file_path: Path):
        """Export data as JSON"""
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str, ensure_ascii=False)
    
    async def _export_csv(self, data: Any, file_path: Path):
        """Export data as CSV"""
        if isinstance(data, list) and data:
            # Flatten data for CSV export
            flattened_data = []
            for item in data:
                if isinstance(item, dict):
                    flat_item = self._flatten_dict(item)
                    flattened_data.append(flat_item)
                else:
                    flattened_data.append({'value': str(item)})
            
            if flattened_data:
                fieldnames = set()
                for item in flattened_data:
                    fieldnames.update(item.keys())
                
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=list(fieldnames))
                    writer.writeheader()
                    writer.writerows(flattened_data)
        else:
            # Simple data
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                if isinstance(data, dict):
                    for key, value in data.items():
                        writer.writerow([key, value])
                else:
                    writer.writerow([str(data)])
    
    async def _export_yaml(self, data: Any, file_path: Path):
        """Export data as YAML"""
        with open(file_path, 'w', encoding='utf-8') as f:
            yaml.dump(data, f, default_flow_style=False, indent=2, allow_unicode=True)
    
    async def _export_html(self, data: Any, file_path: Path):
        """Export data as HTML"""
        html_content = await self._generate_html_export(data)
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
    
    async def _export_xml(self, data: Any, file_path: Path):
        """Export data as XML"""
        xml_content = self._dict_to_xml(data, 'wwyvq_export')
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
            f.write(xml_content)
    
    async def _export_excel(self, data: Any, file_path: Path):
        """Export data as Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "WWYVQ Export"
            
            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            if isinstance(data, list) and data:
                # Flatten data for Excel export
                flattened_data = []
                for item in data:
                    if isinstance(item, dict):
                        flat_item = self._flatten_dict(item)
                        flattened_data.append(flat_item)
                    else:
                        flattened_data.append({'value': str(item)})
                
                if flattened_data:
                    # Write headers
                    headers = list(flattened_data[0].keys())
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                    
                    # Write data
                    for row, item in enumerate(flattened_data, 2):
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=row, column=col, value=item.get(header, ''))
            
            wb.save(file_path)
            
        except ImportError:
            # Fallback to CSV if openpyxl not available
            csv_path = file_path.with_suffix('.csv')
            await self._export_csv(data, csv_path)
            # Rename to xlsx for consistency
            csv_path.rename(file_path)
    
    def _flatten_dict(self, d: Dict, parent_key: str = '', sep: str = '_') -> Dict:
        """Flatten nested dictionary"""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                items.append((new_key, ', '.join(str(x) for x in v)))
            else:
                items.append((new_key, v))
        return dict(items)
    
    def _dict_to_xml(self, data: Any, root_name: str = 'root') -> str:
        """Convert dictionary to XML"""
        def dict_to_xml_recursive(d, name):
            if isinstance(d, dict):
                xml = f"<{name}>"
                for key, value in d.items():
                    xml += dict_to_xml_recursive(value, key)
                xml += f"</{name}>"
                return xml
            elif isinstance(d, list):
                xml = f"<{name}>"
                for item in d:
                    xml += dict_to_xml_recursive(item, 'item')
                xml += f"</{name}>"
                return xml
            else:
                return f"<{name}>{str(d)}</{name}>"
        
        return dict_to_xml_recursive(data, root_name)
    
    def _count_records(self, data: Any) -> int:
        """Count records in data"""
        if isinstance(data, list):
            return len(data)
        elif isinstance(data, dict):
            # Count top-level keys
            return len(data)
        else:
            return 1
    
    def _calculate_checksum(self, file_path: Path) -> str:
        """Calculate file checksum"""
        hasher = hashlib.md5()
        with open(file_path, 'rb') as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hasher.update(chunk)
        return hasher.hexdigest()
    
    async def _generate_html_export(self, data: Any) -> str:
        """Generate HTML export"""
        html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>WWYVQ v2.1 - Data Export</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .header {{ text-align: center; margin-bottom: 30px; }}
        .header h1 {{ color: #333; margin: 0; }}
        .header p {{ color: #666; margin: 5px 0; }}
        .data-section {{ margin: 20px 0; }}
        .data-section h2 {{ color: #444; border-bottom: 2px solid #007acc; padding-bottom: 5px; }}
        pre {{ background: #f8f9fa; padding: 15px; border-radius: 5px; overflow-x: auto; }}
        .timestamp {{ text-align: right; color: #888; font-size: 12px; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🔥 WWYVQ Framework v2.1</h1>
            <p>Professional Data Export</p>
            <div class="timestamp">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}</div>
        </div>
        
        <div class="data-section">
            <h2>📊 Exported Data</h2>
            <pre>{json.dumps(data, indent=2, default=str, ensure_ascii=False)}</pre>
        </div>
    </div>
</body>
</html>
"""
        return html
    
    async def _generate_summary_report(self, data: Dict[str, Any]) -> str:
        """Generate summary report"""
        stats = data.get('statistics', {})
        operations = data.get('operations', [])
        
        html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>WWYVQ v2.1 - Summary Report</title>
    <style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 0; padding: 20px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; border-radius: 12px; box-shadow: 0 8px 32px rgba(0,0,0,0.1); overflow: hidden; }}
        .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; text-align: center; }}
        .header h1 {{ margin: 0; font-size: 2.5em; }}
        .header p {{ margin: 10px 0 0; opacity: 0.9; }}
        .content {{ padding: 30px; }}
        .stats-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 30px 0; }}
        .stat-card {{ background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); color: white; padding: 20px; border-radius: 8px; text-align: center; }}
        .stat-card h3 {{ margin: 0 0 10px; font-size: 1.8em; }}
        .stat-card p {{ margin: 0; opacity: 0.9; }}
        .section {{ margin: 30px 0; }}
        .section h2 {{ color: #333; border-bottom: 3px solid #667eea; padding-bottom: 10px; }}
        .operations-list {{ background: #f8f9fa; padding: 20px; border-radius: 8px; }}
        .operation-item {{ background: white; margin: 10px 0; padding: 15px; border-radius: 6px; border-left: 4px solid #667eea; }}
        .timestamp {{ text-align: right; color: #888; font-size: 12px; margin-top: 30px; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🔥 WWYVQ Framework v2.1</h1>
            <p>Executive Summary Report</p>
        </div>
        
        <div class="content">
            <div class="stats-grid">
                <div class="stat-card">
                    <h3>{stats.get('operations_total', 0)}</h3>
                    <p>Total Operations</p>
                </div>
                <div class="stat-card">
                    <h3>{stats.get('targets_processed', 0)}</h3>
                    <p>Targets Processed</p>
                </div>
                <div class="stat-card">
                    <h3>{stats.get('credentials_found', 0)}</h3>
                    <p>Credentials Found</p>
                </div>
                <div class="stat-card">
                    <h3>{stats.get('valid_credentials', 0)}</h3>
                    <p>Valid Credentials</p>
                </div>
            </div>
            
            <div class="section">
                <h2>📊 Recent Operations</h2>
                <div class="operations-list">
                    {"".join([f'<div class="operation-item"><strong>{op.get("type", "Unknown")}</strong> - {op.get("targets", 0)} targets processed</div>' for op in operations[-5:]])}
                </div>
            </div>
            
            <div class="timestamp">
                Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}
            </div>
        </div>
    </div>
</body>
</html>
"""
        return html
    
    async def _generate_detailed_report(self, data: Dict[str, Any]) -> str:
        """Generate detailed technical report"""
        # Implementation would include detailed charts, graphs, and comprehensive data
        return await self._generate_summary_report(data)  # Simplified for now
    
    async def _generate_executive_report(self, data: Dict[str, Any]) -> str:
        """Generate executive-level report"""
        # Implementation would focus on high-level metrics and business impact
        return await self._generate_summary_report(data)  # Simplified for now
    
    async def _generate_technical_report(self, data: Dict[str, Any]) -> str:
        """Generate technical report"""
        # Implementation would include technical details, configurations, and system info
        return await self._generate_summary_report(data)  # Simplified for now
    
    async def _generate_credentials_report(self, data: Dict[str, Any]) -> str:
        """Generate credentials-specific report"""
        # Implementation would focus on credential analysis and validation results
        return await self._generate_summary_report(data)  # Simplified for now
    
    async def _generate_statistics_report(self, data: Dict[str, Any]) -> str:
        """Generate statistics report"""
        # Implementation would include detailed charts and statistical analysis
        return await self._generate_summary_report(data)  # Simplified for now
    
    async def _get_session_data(self, session_id: str) -> Optional[Dict[str, Any]]:
        """Get session data from session manager"""
        try:
            if hasattr(self.engine, 'session_manager'):
                session = await self.engine.session_manager.get_session(session_id)
                if session:
                    return asdict(session)
        except Exception as e:
            self.logger.error(f"Failed to get session data: {e}")
        return None
    
    async def _get_session_results(self, session_id: str) -> Optional[Dict[str, Any]]:
        """Get session results"""
        # This would integrate with results storage
        return {'results': [], 'session_id': session_id}
    
    async def _get_session_logs(self, session_id: str) -> Optional[List[Dict[str, Any]]]:
        """Get session logs"""
        try:
            if hasattr(self.engine, 'logger'):
                logs = self.engine.logger.search_logs(session_id=session_id, limit=1000)
                return [log.to_dict() for log in logs]
        except Exception as e:
            self.logger.error(f"Failed to get session logs: {e}")
        return []
    
    async def _add_session_reports(self, zipf: zipfile.ZipFile, session_id: str):
        """Add session reports to archive"""
        # Add any existing reports for this session
        session_reports = list(self.reports_dir.glob(f"*{session_id}*"))
        for report_path in session_reports:
            zipf.write(report_path, f"reports/{report_path.name}")
    
    async def cleanup_old_exports(self, days: int = 30):
        """Clean up old export files"""
        try:
            cutoff_date = datetime.now() - timedelta(days=days)
            
            cleaned_count = 0
            for directory in [self.exports_dir, self.reports_dir, self.archives_dir]:
                for file_path in directory.iterdir():
                    if file_path.is_file():
                        file_time = datetime.fromtimestamp(file_path.stat().st_mtime)
                        if file_time < cutoff_date:
                            file_path.unlink()
                            cleaned_count += 1
            
            self.logger.info(f"🧹 Cleaned up {cleaned_count} old export files")
            
        except Exception as e:
            self.logger.error(f"Failed to cleanup old exports: {e}")
    
    async def get_statistics(self) -> Dict[str, Any]:
        """Get export statistics"""
        stats = self.stats.copy()
        
        # Add directory sizes
        stats['export_dir_size'] = sum(f.stat().st_size for f in self.exports_dir.rglob('*') if f.is_file())
        stats['reports_dir_size'] = sum(f.stat().st_size for f in self.reports_dir.rglob('*') if f.is_file())
        stats['archives_dir_size'] = sum(f.stat().st_size for f in self.archives_dir.rglob('*') if f.is_file())
        
        return stats
    
    async def shutdown(self):
        """Shutdown module"""
        self.logger.info("🛑 Data Export & Reporting Module shutdown completed")